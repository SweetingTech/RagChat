from PIL import Image
import pytesseract  # Ensure Tesseract OCR is installed
import pypdf
import docx
import openpyxl
from bs4 import BeautifulSoup
import requests
import os

def load_document(file_path):
    """Load a document from a file path and return its text content."""
    file_extension = os.path.splitext(file_path)[1].lower()
    try:
        if file_extension == ".pdf":
            return load_pdf(file_path)
        elif file_extension == ".docx":
            return load_docx(file_path)
        elif file_extension == ".txt":
            return load_txt(file_path)
        elif file_extension == ".xlsx":
            return load_xlsx(file_path)
        elif file_extension in (".jpg", ".jpeg", ".png", ".gif"):
            return load_image(file_path)
        elif file_extension == ".html":
            return load_html(file_path)
        else:
            print(f"Unsupported file type: {file_extension} - attempting basic text extraction")
            try:
                with open(file_path, "r", encoding="utf-8") as f:
                    return f.read()
            except Exception as e:
                print(f"Error reading file: {e}")
                return None
    except Exception as e:
        print(f"Error loading document {file_path}: {e}")
        return None

def load_pdf(file_path):
    """Extract text from a PDF file."""
    text = ""
    try:
        with open(file_path, "rb") as f:
            reader = pypdf.PdfReader(f)
            for page in reader.pages:
                text += page.extract_text()
    except Exception as e:
        print(f"Error loading PDF {file_path}: {e}")
        return None
    return text

def load_docx(file_path):
    """Extract text from a DOCX file."""
    try:
        doc = docx.Document(file_path)
        return "\n".join([paragraph.text for paragraph in doc.paragraphs])
    except Exception as e:
        print(f"Error loading DOCX {file_path}: {e}")
        return None

def load_txt(file_path):
    """Extract text from a TXT file."""
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            return f.read()
    except Exception as e:
        print(f"Error loading TXT {file_path}: {e}")
        return None

def load_xlsx(file_path):
    """Extract text from an XLSX file."""
    try:
        workbook = openpyxl.load_workbook(file_path)
        text = ""
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            for row in worksheet.iter_rows():
                text += " ".join(str(cell.value) for cell in row if cell.value is not None) + "\\n"
        return text
    except Exception as e:
        print(f"Error loading XLSX {file_path}: {e}")
        return None

def load_image(file_path):
    """Extract text from an image using OCR."""
    try:
        img = Image.open(file_path)
        return pytesseract.image_to_string(img)
    except Exception as e:
        print(f"Error loading image {file_path} or running OCR: {e}")
        return None

def load_html(file_path):
    """Extract text from an HTML file."""
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            soup = BeautifulSoup(f, "html.parser")
            return soup.get_text()
    except Exception as e:
        print(f"Error loading HTML {file_path}: {e}")
        return None
